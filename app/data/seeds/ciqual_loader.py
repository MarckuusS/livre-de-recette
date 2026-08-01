"""Idempotent loader for the CIQUAL (ANSES) nutrition table.

The CIQUAL file is *not* shipped with the repo (license + size). The user downloads it
manually from https://ciqual.anses.fr/ and saves it under `app/data/seeds/`. Both .xls
(historical CIQUAL releases) and .csv exports are accepted — the loader auto-detects
the format from the file extension.

Run as a script:
    python -m app.data.seeds.ciqual_loader [path/to/ciqual.xls|.csv]

The loader is idempotent: rerunning updates rows in place via
IngredientRepo.upsert_by_source_ref (matched on source='ciqual', source_ref=alim_code).
"""

from __future__ import annotations

import csv
import logging
import re
import sys
from collections.abc import Iterable
from pathlib import Path

from app.data.db import init_schema, make_engine, make_session_factory, session_scope
from app.data.repositories import IngredientRepo
from app.domain.models import Ingredient, Source

log = logging.getLogger(__name__)

# Column-name candidates (CIQUAL has slightly different headers between versions and
# between the .xls and .csv exports). All keys are normalized via `_norm` before lookup
# (lowercase + strip + collapse whitespace + drop diacritics-insensitive characters).
_COLUMN_CANDIDATES: dict[str, list[str]] = {
    "code": ["alim_code", "code"],
    "name": ["alim_nom_fr", "nom", "name"],
    # CIQUAL 2025 .xls uses '(kcal 100 g)' (no slash); historical .csv uses '(kcal/100 g)'.
    # We normalize '/' to ' ' in `_norm`, so both shapes match the candidates below.
    "kcal": [
        "energie, reglement ue n 1169 2011 (kcal 100 g)",
        "energie, reglement ue n° 1169 2011 (kcal 100 g)",
        "energie (kcal 100 g)",
        "energie, n x facteur jones, avec fibres (kcal 100 g)",
        "kcal 100g",
    ],
    "proteins": [
        "proteines, n x facteur de jones (g 100 g)",
        "proteines (g 100 g)",
    ],
    "carbs": ["glucides (g 100 g)", "glucides"],
    "sugars": ["sucres (g 100 g)", "sucres"],
    "fats": ["lipides (g 100 g)", "lipides"],
    "saturated_fats": [
        "ag satures (g 100 g)",
        "acides gras satures (g 100 g)",
    ],
    "fiber": [
        "fibres alimentaires (g 100 g)",
        "fibres (g 100 g)",
    ],
    "salt": [
        "sel chlorure de sodium (g 100 g)",
        "sel (g 100 g)",
    ],
    "category_l1": ["alim_grp_nom_fr"],
    "category_l2": ["alim_ssgrp_nom_fr"],
}

# Folded diacritics so 'protéines' and 'proteines' both normalize to 'proteines'.
_FOLD_MAP = str.maketrans({
    "à": "a", "á": "a", "â": "a", "ä": "a", "ã": "a",
    "é": "e", "è": "e", "ê": "e", "ë": "e",
    "í": "i", "ì": "i", "î": "i", "ï": "i",
    "ó": "o", "ò": "o", "ô": "o", "ö": "o", "õ": "o",
    "ú": "u", "ù": "u", "û": "u", "ü": "u",
    "ý": "y", "ÿ": "y",
    "ç": "c",
})


def _norm(s: str) -> str:
    """Normalize a header for matching: lowercase, fold accents, slash -> space, collapse spaces."""
    s = s.strip().lower().translate(_FOLD_MAP)
    # Treat slash, line-break and tab as whitespace (CIQUAL .xls headers contain '\n').
    s = re.sub(r"[/\n\t]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _parse_float(raw) -> float | None:
    """Parse a CIQUAL numeric cell. Handles '', '-', 'traces', '< 0.1', decimal commas, and floats."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        f = float(raw)
        # xls sometimes encodes 'empty' as 0 — we can't tell apart. Treat exact 0 as None
        # only for the energy column? No: 0 is meaningful (water has 0 kcal). Keep it.
        return f
    s = str(raw).strip()
    if s in ("", "-"):
        return None
    if s.lower() in ("traces", "trace"):
        return 0.0
    if s.startswith("<"):
        # '< 0.1' -> treat as 0 conservatively (avoids inflating totals).
        return 0.0
    s = s.replace(",", ".").replace(" ", "").replace("\xa0", "")
    try:
        return float(s)
    except ValueError:
        log.debug("Could not parse CIQUAL cell %r", raw)
        return None


def _resolve_columns(headers: list[str]) -> dict[str, str | None]:
    """Map our logical column keys to the actual headers found in the file."""
    norm_to_orig: dict[str, str] = {_norm(h): h for h in headers}
    resolved: dict[str, str | None] = {}
    for logical, candidates in _COLUMN_CANDIDATES.items():
        wanted = {_norm(c) for c in candidates}
        match = next((orig for n, orig in norm_to_orig.items() if n in wanted), None)
        resolved[logical] = match
    return resolved


# --------------------------------------------------------------------------- #
# CSV reader
# --------------------------------------------------------------------------- #


def _read_csv(path: Path) -> Iterable[dict[str, str]]:
    """Read CSV trying utf-8 then latin-1; CIQUAL CSV uses ';' separator."""
    encodings = ("utf-8-sig", "utf-8", "latin-1")
    for enc in encodings:
        try:
            with path.open("r", encoding=enc, newline="") as fh:
                sample = fh.read(2048)
                fh.seek(0)
                dialect = csv.Sniffer().sniff(sample, delimiters=";,")
                reader = csv.DictReader(fh, dialect=dialect)
                yield from reader
            return
        except (UnicodeDecodeError, csv.Error):
            continue
    raise RuntimeError(f"Could not read CSV at {path} with any of {encodings}")


# --------------------------------------------------------------------------- #
# XLS / XLSX reader
# --------------------------------------------------------------------------- #


def _read_xls(path: Path) -> Iterable[dict]:
    """Read an .xls (binary OLE2) file via xlrd. Returns dicts keyed by header strings."""
    try:
        import xlrd  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError(
            "Reading .xls requires xlrd<2.0. Install it with:\n"
            "    .venv/Scripts/pip install 'xlrd<2.0'"
        ) from exc

    book = xlrd.open_workbook(str(path))
    # CIQUAL workbook has 'composition nutritionnelle' as sheet 0.
    sheet = book.sheet_by_index(0)
    if sheet.nrows < 2:
        return
    headers = [str(h) for h in sheet.row_values(0)]
    for r in range(1, sheet.nrows):
        values = sheet.row_values(r)
        yield dict(zip(headers, values))


def _read_xlsx(path: Path) -> Iterable[dict]:
    """Read an .xlsx file via openpyxl. Same shape as _read_xls."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError(
            "Reading .xlsx requires openpyxl. Install it with:\n"
            "    .venv/Scripts/pip install openpyxl"
        ) from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows)]
    for values in rows:
        yield dict(zip(headers, values))


def _read_any(path: Path) -> Iterable[dict]:
    """Dispatch on file extension."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xls":
        return _read_xls(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise RuntimeError(f"Unsupported CIQUAL file extension {suffix!r}: expected .csv, .xls, or .xlsx")


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #


def load_csv(path: Path) -> int:
    """Load (or update) all ingredients from the CIQUAL file. Name kept for API compat;
    the file may actually be .xls / .xlsx — auto-detected by extension.
    """
    if not path.exists():
        log.warning("CIQUAL file not found at %s — skipping seed.", path)
        return 0

    engine = make_engine()
    init_schema(engine)
    factory = make_session_factory(engine)

    rows = list(_read_any(path))
    if not rows:
        log.warning("CIQUAL file %s is empty.", path)
        return 0

    headers = list(rows[0].keys())
    cols = _resolve_columns(headers)
    if cols["code"] is None or cols["name"] is None:
        raise RuntimeError(
            f"CIQUAL file missing required columns 'alim_code'/'alim_nom_fr'. "
            f"Headers seen: {headers[:10]}..."
        )

    count = 0
    with session_scope(factory) as session:
        repo = IngredientRepo(session)
        for row in rows:
            code = str(row.get(cols["code"]) or "").strip()
            name = str(row.get(cols["name"]) or "").strip()
            if not code or not name:
                continue
            # CIQUAL .xls stores numeric alim_code as a float ('20055.0') — normalize.
            if code.endswith(".0"):
                code = code[:-2]

            cat_l1 = (str(row.get(cols["category_l1"]) or "").strip() if cols["category_l1"] else "")
            cat_l2 = (str(row.get(cols["category_l2"]) or "").strip() if cols["category_l2"] else "")
            ing = Ingredient(
                name=name,
                source=Source.CIQUAL,
                source_ref=code,
                kcal_per_100g=_parse_float(row.get(cols["kcal"])) if cols["kcal"] else None,
                proteins_g=_parse_float(row.get(cols["proteins"])) if cols["proteins"] else None,
                carbs_g=_parse_float(row.get(cols["carbs"])) if cols["carbs"] else None,
                sugars_g=_parse_float(row.get(cols["sugars"])) if cols["sugars"] else None,
                fats_g=_parse_float(row.get(cols["fats"])) if cols["fats"] else None,
                saturated_fats_g=_parse_float(row.get(cols["saturated_fats"])) if cols["saturated_fats"] else None,
                fiber_g=_parse_float(row.get(cols["fiber"])) if cols["fiber"] else None,
                salt_g=_parse_float(row.get(cols["salt"])) if cols["salt"] else None,
                category_l1=cat_l1 or None,
                category_l2=cat_l2 or None,
                in_personal_library=False,  # CIQUAL is a source catalog, not the user's library
            )
            # Don't clobber `in_personal_library` if the user has already imported this row
            # — only update the nutrition data and keep their flag.
            existing = repo.find_by_source_ref(Source.CIQUAL, code)
            if existing is not None and existing.in_personal_library:
                ing = ing.model_copy(update={"in_personal_library": True})
            repo.upsert_by_source_ref(ing)
            count += 1
    log.info("CIQUAL: processed %d rows from %s", count, path)
    return count


def default_csv_path() -> Path:
    """Where the loader looks for the CIQUAL file by default. Tries .xls first, then .xlsx, then .csv."""
    base = Path(__file__).parent
    for ext in ("xls", "xlsx", "csv"):
        candidate = base / f"ciqual.{ext}"
        if candidate.exists():
            return candidate
    return base / "ciqual.csv"


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    argv = argv or sys.argv[1:]
    path = Path(argv[0]) if argv else default_csv_path()
    n = load_csv(path)
    print(f"CIQUAL loader: {n} rows processed from {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
